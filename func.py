import os
import pandas as pd
import openpyxl

import path

def open_template(template):
    wb = openpyxl.load_workbook(template, read_only=False, keep_vba=True)
    return wb

def open_table(table):
    table = pd.read_excel(table, sheet_name='clc_data_lukoil_2Q_23')
    return table

def open_origin_clc(path, sheet):
    print("start open old_clc")
    wb = openpyxl.load_workbook(path, read_only=True, data_only = True)
    ws = wb[sheet]
    print("old clc opened")
    return ws

def fill_payment(dic):
    print(dic['vehicle_cost'])
    print(type(dic['vehicle_cost']))
    agr_no = dic['agr_no']
    vehicle_cost = dic['vehicle_cost']
    file_name = dic['file_name']
    age = dic['age']
    rv_net = dic['rv_net']
    smr_diff = dic['smr_diff']
    tire_diff = dic['tire_diff']
    tire_storage = dic['tire_storage']
    tire_change = dic['tire_change']
    pzu = dic['pzu']
    registration = dic['registration']
    rsa = dic['rsa']
    telem_install = dic['telem_install']
    telem_service = dic['telem_service']
    kasko = dic['kasko']
    osago = dic['osago']
    dsago = dic['dsago']
    tax = dic['tax']
    margin = dic['margin']
    ets = dic['ets']
    return_fee = dic['return_fee']
    credit_remain = dic['credit_remain']


    template = open_template(path.template)
    ws = template['payment']
    ws['C10'] = vehicle_cost
    if age <= 12:
        ws['D17'] = rv_net
    elif age <= 24:
        ws['D18'] = rv_net
    elif age <= 36:
        ws['D19'] = rv_net
    elif age <= 48:
        ws['D20'] = rv_net
    elif age <= 60:
        ws['D21'] = rv_net
    else:
        print('check the age')

    ws['C30'] = ets
    ws['C31'] = margin
    ws['C32'] = pzu
    ws['C35'] = age
    ws['C37'] = smr_diff
    ws['C40'] = tire_diff
    ws['C41'] = tire_storage
    ws['C42'] = tire_change
    ws['C44'] = registration
    ws['C45'] = rsa
    ws['C46'] = telem_install
    ws['C47'] = telem_service
    ws['C52'] = return_fee
    ws['C55'] = kasko
    ws['C63'] = osago
    ws['C71'] = dsago
    ws['C79'] = tax
    print("payment sheet completed")

    # __________________________
    ws = template['PNL_new']
    ws['H110'] = credit_remain
    ws['H112'] = credit_remain

    print("PNL sheet completed")

    ws = template['original']
    old = open_origin_clc(path.origin_clc2, agr_no)
    counter = 0
    total_cells = 130*120
    for i in range(1, 10):  # должно быть 130
        for j in range(1, 10):  # должно быть 120
            # reading cell value from source excel file
            c = old.cell(row=i, column=j)
            #print(str(c) + "is copied")

            # writing the read value to destination excel file
            ws.cell(row=i, column=j).value = c.value
            counter +=1

            print(str(counter) + " is pasted from " + str(total_cells))


    file = str(file_name) + ' clc_recontracting ' + '.xlsm'
    savename = os.path.join(path.savepath, file)
    template.save(savename)
    print(str(agr_no) + " clc saved")
    return 0

